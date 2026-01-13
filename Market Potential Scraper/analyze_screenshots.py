"""
分析 screenshots 目录下的 DataEye API JSON 数据
生成维度分析表格
"""

import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Tuple
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ScreenshotAnalyzer:
    """分析 screenshots 目录下的 JSON 数据"""

    def __init__(self, screenshots_dir: str = "screenshots"):
        """
        初始化分析器

        Args:
            screenshots_dir: screenshots 目录路径
        """
        self.screenshots_dir = screenshots_dir
        self.drama_data = {}  # 存储每个剧目的聚合数据

    def load_all_json_files(self) -> Dict[str, List[Dict]]:
        """
        加载所有 JSON 文件

        Returns:
            字典: {剧名: [2年数据, 30天数据]}
        """
        logger.info(f"开始扫描目录: {self.screenshots_dir}")

        if not os.path.exists(self.screenshots_dir):
            logger.error(f"目录不存在: {self.screenshots_dir}")
            return {}

        # 按剧名分组
        drama_files = defaultdict(dict)

        for filename in os.listdir(self.screenshots_dir):
            if not filename.endswith('.json'):
                continue

            # 解析文件名: 剧名_时间窗口.json
            if '_2年' in filename:
                drama_name = filename.replace('_2年.json', '')
                time_window = '2年'
            elif '_30天' in filename:
                drama_name = filename.replace('_30天.json', '')
                time_window = '30天'
            else:
                logger.warning(f"无法识别文件名格式: {filename}")
                continue

            filepath = os.path.join(self.screenshots_dir, filename)

            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    drama_files[drama_name][time_window] = data
                    logger.info(f"✓ 加载: {filename}")
            except Exception as e:
                logger.error(f"✗ 加载失败 {filename}: {e}")

        logger.info(f"✓ 共加载 {len(drama_files)} 个剧目的数据")
        return dict(drama_files)

    def extract_dimensions(self, drama_name: str, data_2y: Dict, data_30d: Dict) -> Dict:
        """
        提取维度数据

        Args:
            drama_name: 剧名
            data_2y: 2年数据
            data_30d: 30天数据

        Returns:
            维度数据字典
        """
        logger.info(f"分析剧目: {drama_name}")

        # 提取素材列表
        materials_2y = data_2y.get('content', {}).get('searchList', [])
        materials_30d = data_30d.get('content', {}).get('searchList', []) if data_30d else []

        # 初始化维度数据
        dimensions = {
            '剧名': drama_name,

            # 投放规模
            '2年累计曝光量': 0,
            '2年预估下载量': 0,
            '30天累计曝光量': 0,
            '30天预估下载量': 0,

            # 素材数
            '2年素材总数': len(materials_2y),
            '30天在投素材数': len(materials_30d),

            # 地区分布
            '投放国家数_2年': 0,
            '投放国家数_30天': 0,
            'Top5国家_2年': [],
            'Top5国家_30天': [],

            # 平台分布
            '投放平台分布_2年': {},
            '投放平台分布_30天': {},

            # 素材形式
            '素材形式占比_2年': {},
            '素材形式占比_30天': {},

            # 制作方
            '主要制作方_2年': [],
            '主要制作方_30天': [],

            # 生命周期
            '首次投放日期': None,
            '最后投放日期': None,
            '生命周期天数': 0,
            '活跃天数': 0,
        }

        # 分析 2年数据
        if materials_2y:
            dimensions.update(self._analyze_materials(materials_2y, '2年'))

        # 分析 30天数据
        if materials_30d:
            dimensions.update(self._analyze_materials(materials_30d, '30天'))

        return dimensions

    def _analyze_materials(self, materials: List[Dict], time_window: str) -> Dict:
        """
        分析素材列表，提取维度数据

        Args:
            materials: 素材列表
            time_window: 时间窗口 ('2年' 或 '30天')

        Returns:
            维度数据字典
        """
        result = {}

        # 统计变量
        total_exposure = 0
        total_downloads = 0
        country_stats = defaultdict(int)  # 国家 -> 曝光量
        platform_stats = defaultdict(int)  # 平台 -> 素材数
        material_type_stats = defaultdict(int)  # 素材类型 -> 素材数
        publisher_stats = defaultdict(int)  # 制作方 -> 素材数
        first_seen_dates = []
        last_seen_dates = []
        active_days_set = set()

        # 遍历素材
        for material in materials:
            # 曝光量和下载量
            exposure = material.get('exposureNum', 0)
            downloads = material.get('downloadNum', 0)
            total_exposure += exposure
            total_downloads += downloads

            # 国家统计
            countries = material.get('countries', [])
            for country in countries:
                country_name = country.get('countryName', '未知')
                country_stats[country_name] += exposure

            # 平台统计
            media = material.get('media', {})
            if media:
                platform_name = media.get('mediaName', '未知')
                platform_stats[platform_name] += 1

            # 素材类型统计
            material_type = material.get('materialType', 0)
            type_name = '视频' if material_type == 2 else '图片'
            material_type_stats[type_name] += 1

            # 制作方统计
            publisher = material.get('publisher', {})
            if publisher:
                publisher_name = publisher.get('publisherName', '未知')
                publisher_stats[publisher_name] += 1

            # 时间统计
            first_seen = material.get('firstSeen', '')
            last_seen = material.get('lastSeen', '')
            if first_seen:
                first_seen_dates.append(first_seen)
            if last_seen:
                last_seen_dates.append(last_seen)

            # 活跃天数
            release_days = material.get('releaseDay', 0)
            if first_seen and release_days > 0:
                try:
                    from datetime import datetime, timedelta
                    start_date = datetime.strptime(first_seen, '%Y-%m-%d')
                    for i in range(release_days):
                        day = start_date + timedelta(days=i)
                        active_days_set.add(day.strftime('%Y-%m-%d'))
                except:
                    pass

        # 汇总结果
        suffix = f'_{time_window}'

        result[f'累计曝光量{suffix}'] = total_exposure
        result[f'预估下载量{suffix}'] = total_downloads

        # Top 5 国家
        top_countries = sorted(country_stats.items(), key=lambda x: x[1], reverse=True)[:5]
        result[f'投放国家数{suffix}'] = len(country_stats)
        result[f'Top5国家{suffix}'] = [f"{c[0]}({c[1]:,})" for c in top_countries]

        # 平台分布
        result[f'投放平台分布{suffix}'] = dict(platform_stats)

        # 素材形式占比
        result[f'素材形式占比{suffix}'] = dict(material_type_stats)

        # Top 3 制作方
        top_publishers = sorted(publisher_stats.items(), key=lambda x: x[1], reverse=True)[:3]
        result[f'主要制作方{suffix}'] = [f"{p[0]}({p[1]})" for p in top_publishers]

        # 生命周期（仅2年数据计算）
        if time_window == '2年' and first_seen_dates and last_seen_dates:
            result['首次投放日期'] = min(first_seen_dates)
            result['最后投放日期'] = max(last_seen_dates)
            try:
                from datetime import datetime
                first = datetime.strptime(min(first_seen_dates), '%Y-%m-%d')
                last = datetime.strptime(max(last_seen_dates), '%Y-%m-%d')
                result['生命周期天数'] = (last - first).days
                result['活跃天数'] = len(active_days_set)
            except:
                pass

        return result

    def generate_excel_report(self, output_file: str = "维度分析报告.xlsx"):
        """
        生成 Excel 分析报告

        Args:
            output_file: 输出文件名
        """
        logger.info(f"开始生成 Excel 报告: {output_file}")

        # 加载所有 JSON 数据
        drama_files = self.load_all_json_files()

        if not drama_files:
            logger.error("没有找到任何数据文件")
            return

        # 提取所有剧目的维度数据
        all_dimensions = []
        for drama_name, files in drama_files.items():
            data_2y = files.get('2年')
            data_30d = files.get('30天')

            if not data_2y:
                logger.warning(f"剧目 {drama_name} 缺少2年数据，跳过")
                continue

            dimensions = self.extract_dimensions(drama_name, data_2y, data_30d)
            all_dimensions.append(dimensions)

        if not all_dimensions:
            logger.error("没有提取到任何维度数据")
            return

        # 创建 Excel 工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认 sheet

        # Sheet 1: 维度汇总表
        self._create_summary_sheet(wb, all_dimensions)

        # Sheet 2: 详细数据明细
        self._create_detail_sheet(wb, all_dimensions)

        # Sheet 3: 制作方分析
        self._create_publisher_sheet(wb, all_dimensions)

        # 保存文件
        wb.save(output_file)
        logger.info(f"✓ Excel 报告已生成: {output_file}")

    def _create_summary_sheet(self, wb: Workbook, all_dimensions: List[Dict]):
        """创建维度汇总表"""
        ws = wb.create_sheet("维度汇总表", 0)

        # 定义表头
        headers = [
            '剧名',
            '2年累计曝光量',
            '30天累计曝光量',
            '2年素材总数',
            '30天在投素材数',
            '生命周期天数',
            '活跃天数',
            '投放国家数_2年',
            'Top5国家_2年',
            '主要制作方_2年',
            '首次投放日期',
            '最后投放日期',
        ]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row_idx, dim in enumerate(all_dimensions, 2):
            ws.cell(row=row_idx, column=1, value=dim.get('剧名', ''))
            ws.cell(row=row_idx, column=2, value=dim.get('累计曝光量_2年', 0))
            ws.cell(row=row_idx, column=3, value=dim.get('累计曝光量_30天', 0))
            ws.cell(row=row_idx, column=4, value=dim.get('2年素材总数', 0))
            ws.cell(row=row_idx, column=5, value=dim.get('30天在投素材数', 0))
            ws.cell(row=row_idx, column=6, value=dim.get('生命周期天数', 0))
            ws.cell(row=row_idx, column=7, value=dim.get('活跃天数', 0))
            ws.cell(row=row_idx, column=8, value=dim.get('投放国家数_2年', 0))
            ws.cell(row=row_idx, column=9, value='\n'.join(dim.get('Top5国家_2年', [])))
            ws.cell(row=row_idx, column=10, value='\n'.join(dim.get('主要制作方_2年', [])))
            ws.cell(row=row_idx, column=11, value=dim.get('首次投放日期', ''))
            ws.cell(row=row_idx, column=12, value=dim.get('最后投放日期', ''))

        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 35
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15

        logger.info("✓ 已创建维度汇总表")

    def _create_detail_sheet(self, wb: Workbook, all_dimensions: List[Dict]):
        """创建详细数据明细表"""
        ws = wb.create_sheet("详细数据明细", 1)

        # 定义表头
        headers = [
            '剧名',
            '2年累计曝光量',
            '2年预估下载量',
            '2年素材总数',
            '30天累计曝光量',
            '30天预估下载量',
            '30天在投素材数',
            '生命周期天数',
            '活跃天数',
            '首次投放日期',
            '最后投放日期',
            '投放国家数_2年',
            '投放国家数_30天',
            'Top5国家_2年',
            'Top5国家_30天',
            '投放平台_2年',
            '投放平台_30天',
            '素材形式_2年',
            '素材形式_30天',
        ]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 写入数据
        for row_idx, dim in enumerate(all_dimensions, 2):
            ws.cell(row=row_idx, column=1, value=dim.get('剧名', ''))
            ws.cell(row=row_idx, column=2, value=dim.get('累计曝光量_2年', 0))
            ws.cell(row=row_idx, column=3, value=dim.get('预估下载量_2年', 0))
            ws.cell(row=row_idx, column=4, value=dim.get('2年素材总数', 0))
            ws.cell(row=row_idx, column=5, value=dim.get('累计曝光量_30天', 0))
            ws.cell(row=row_idx, column=6, value=dim.get('预估下载量_30天', 0))
            ws.cell(row=row_idx, column=7, value=dim.get('30天在投素材数', 0))
            ws.cell(row=row_idx, column=8, value=dim.get('生命周期天数', 0))
            ws.cell(row=row_idx, column=9, value=dim.get('活跃天数', 0))
            ws.cell(row=row_idx, column=10, value=dim.get('首次投放日期', ''))
            ws.cell(row=row_idx, column=11, value=dim.get('最后投放日期', ''))
            ws.cell(row=row_idx, column=12, value=dim.get('投放国家数_2年', 0))
            ws.cell(row=row_idx, column=13, value=dim.get('投放国家数_30天', 0))
            ws.cell(row=row_idx, column=14, value='\n'.join(dim.get('Top5国家_2年', [])))
            ws.cell(row=row_idx, column=15, value='\n'.join(dim.get('Top5国家_30天', [])))

            # 平台分布
            platform_2y = dim.get('投放平台分布_2年', {})
            platform_str_2y = '\n'.join([f"{k}: {v}" for k, v in platform_2y.items()])
            ws.cell(row=row_idx, column=16, value=platform_str_2y)

            platform_30d = dim.get('投放平台分布_30天', {})
            platform_str_30d = '\n'.join([f"{k}: {v}" for k, v in platform_30d.items()])
            ws.cell(row=row_idx, column=17, value=platform_str_30d)

            # 素材形式
            material_2y = dim.get('素材形式占比_2年', {})
            material_str_2y = '\n'.join([f"{k}: {v}" for k, v in material_2y.items()])
            ws.cell(row=row_idx, column=18, value=material_str_2y)

            material_30d = dim.get('素材形式占比_30天', {})
            material_str_30d = '\n'.join([f"{k}: {v}" for k, v in material_30d.items()])
            ws.cell(row=row_idx, column=19, value=material_str_30d)

        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['N'].width = 30
        ws.column_dimensions['O'].width = 30
        ws.column_dimensions['P'].width = 25
        ws.column_dimensions['Q'].width = 25
        ws.column_dimensions['R'].width = 20
        ws.column_dimensions['S'].width = 20

        logger.info("✓ 已创建详细数据明细表")

    def _create_publisher_sheet(self, wb: Workbook, all_dimensions: List[Dict]):
        """创建制作方分析表"""
        ws = wb.create_sheet("制作方分析", 2)

        # 收集所有制作方数据
        publisher_drama_map = defaultdict(list)  # 制作方 -> [剧名列表]

        for dim in all_dimensions:
            drama_name = dim.get('剧名', '')
            publishers_2y = dim.get('主要制作方_2年', [])

            for pub_str in publishers_2y:
                # 解析 "制作方名称(素材数)" 格式
                if '(' in pub_str:
                    pub_name = pub_str.split('(')[0]
                    publisher_drama_map[pub_name].append(drama_name)

        # 定义表头
        headers = ['制作方', '投放剧目数', '剧目列表', '决策价值']

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 按剧目数排序
        sorted_publishers = sorted(publisher_drama_map.items(), key=lambda x: len(x[1]), reverse=True)

        # 写入数据
        for row_idx, (publisher, dramas) in enumerate(sorted_publishers, 2):
            ws.cell(row=row_idx, column=1, value=publisher)
            ws.cell(row=row_idx, column=2, value=len(dramas))
            ws.cell(row=row_idx, column=3, value='\n'.join(dramas))

            # 决策价值判断
            decision_value = self._evaluate_publisher(publisher, len(dramas))
            ws.cell(row=row_idx, column=4, value=decision_value)

        # 调整列宽
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50

        logger.info("✓ 已创建制作方分析表")

    def _evaluate_publisher(self, publisher_name: str, drama_count: int) -> str:
        """
        评估制作方的决策价值

        Args:
            publisher_name: 制作方名称
            drama_count: 投放剧目数

        Returns:
            决策价值描述
        """
        # S级制作方（头部玩家）
        s_tier_publishers = ['ReelShort', 'DramaBox', 'FlexTV', 'ShortMax', 'MoboReels']

        for s_pub in s_tier_publishers:
            if s_pub.lower() in publisher_name.lower():
                return f"⭐ S级制作方 - 头部玩家正在投放，验证过的优质剧目"

        # 高频投放方
        if drama_count >= 5:
            return f"🔥 高频投放 - 该制作方投放{drama_count}部剧，可能是专业团队"
        elif drama_count >= 3:
            return f"📊 中等投放 - 该制作方投放{drama_count}部剧"
        else:
            return f"📌 小规模投放 - 该制作方投放{drama_count}部剧"


def main():
    """主函数"""
    logger.info("="*60)
    logger.info("DataEye Screenshots 数据分析工具")
    logger.info("="*60)

    # 创建分析器
    analyzer = ScreenshotAnalyzer(screenshots_dir="screenshots")

    # 生成 Excel 报告
    output_file = f"维度分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    analyzer.generate_excel_report(output_file)

    logger.info("="*60)
    logger.info("分析完成！")
    logger.info("="*60)


if __name__ == "__main__":
    main()
